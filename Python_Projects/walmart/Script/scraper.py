from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException
from bs4 import BeautifulSoup
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import time
import string


class Scraper:
    def __init__(self, driver_path, site_url, readfile_path, writefile_path):
        self.siteUrl = site_url
        self.driverPath = driver_path
        self.readFilePath = readfile_path
        self.writeFilePath = writefile_path
        self.driver = webdriver.Chrome(executable_path=self.driverPath, options=webdriver.ChromeOptions())

    @staticmethod
    def get_states(self):
        path = "../ReadFile/Stores.xlsx"
        wb = load_workbook(path)
        sheet = wb.active
        states = []
        state = ''

        for row in range(2, sheet.max_row):
            if state != sheet.cell(row=row, column=7).value:
                state = sheet.cell(row=row, column=7).value
                states.append(state)
        return states

    def get_aisle(self, path):
        try:
            self.driver.get(path)
        except:
            self.driver.close()

        #     try:
        #         WebDriverWait(self.driver, 3).until(
        #         EC.presence_of_element_located((By.CLASS_NAME, 'nearby-store-count')))
        #     except TimeoutException:
        #         self.driver.quit()
        #         continue
        try:
            search_tab = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'results-info')))
        except NoSuchElementException:
            search_tab = None
        except TimeoutException:
            search_tab = None

        if search_tab is None:
            return None
        soup = BeautifulSoup(self.driver.page_source, 'html.parser')
        div_tags = soup.find('div', attrs={'class': 'search-tab container'}).find_all('div', attrs={'class': 'tile-in-store-info'})
        for div_tag in div_tags:
            str = div_tag.get_text().split('|')

            if len(str) == 2:
                print(str)
                aisle_txt = str[0].split(' ')
                aisle = aisle_txt[1]
                return aisle

    def run(self):
        states = self.get_states(self)
        s_states = ["AK", "ND", "HI", "VT", "RI", "DE"]

        for state in states:
            for s_state in s_states:
                if state == s_state:
                    states.remove(state)
        print(states)

        i = 0
        stores_file = self.readFilePath + '/Stores.xlsx'

        store_wb = load_workbook(stores_file)
        store_sheet = store_wb.active
        k = 1
        for i in range(5, len(states)):
            for store_row in store_sheet:
                if store_row[6].value == states[i]:
                    if k < 13:
                        k += 1
                        continue
                    # make a directory for each state
                    dir_name = "../Result/%s" % (states[i])
                    try:
                        # open(os.path.join(os.pardir, "filename"), "w")
                        os.mkdir(dir_name)
                        print("Directory ", dir_name, " Created ")
                    except FileExistsError:
                        print("Directory ", dir_name, " already exists")

                    store_search_url = self.siteUrl + '/store/' + str(store_row[0].value) + '/' + str(
                        store_row[1].value) + '-ar'
                    items_wb = load_workbook("%s/Items.xlsx" % self.readFilePath)
                    items_sheet = items_wb.active

                    # make a new xlsx file for each store
                    print("File %s-%s.xlsx created" % (store_row[1].value, store_row[0].value))
                    save_path = "%s/%s-%s.xlsx" % (dir_name, store_row[1].value, store_row[0].value)
                    new_wb = Workbook()
                    new_sheet = new_wb.active
                    print(items_sheet.max_row)
                    for item_row in range(1, items_sheet.max_row + 1):
                        item = items_sheet.cell(row=item_row, column=1).value
                        item_search_url = store_search_url + '/search?query=' + item
                        new_sheet.cell(row=item_row, column=1).value = item
                        new_sheet.cell(row=item_row, column=2).value = self.get_aisle(item_search_url)
                        new_wb.save(save_path)
                    # for item_row in items_sheet:
                    #     for cell in item_row:
                    #         item = cell.value
                    #         item_search_url = store_search_url + '/search?query=' + item
                    #         print(item_search_url)
                    #         new_sheet[cell.coordinate].value = cell.value
                    #         coord = 'B%s' % cell.coordinate[1:]
                    #         new_sheet[coord] = self.get_aisle(item_search_url)
                    #         new_wb.save(save_path)
                    new_wb.save(save_path)

        self.driver.close()


if __name__ == '__main__':
    print("================== starting ==================")
    driverPath = "../driver/chromedriver.exe"
    readFilePath = "../ReadFile"
    writeFilePath = "../result"
    siteUrl = "https://www.walmart.com"
    Scraper = Scraper(driverPath, siteUrl, readFilePath, writeFilePath)
    Scraper.run()

    # https://www.walmart.com/

