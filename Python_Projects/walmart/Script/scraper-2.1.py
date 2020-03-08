from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
import csv
from csv import DictWriter
import os
import time
import random


class Scraper:
    def __init__(self, driver_path, site_url, readfile_path, writefile_path):
        self.siteUrl = site_url
        self.driverPath = driver_path
        self.readFilePath = readfile_path
        self.writeFilePath = writefile_path
        self.driver = None
        self.response = False

    @staticmethod
    def get_states():
        f = open("../ReadFile/states.txt", "r")
        states = f.read().split('\n')
        return states

    def init_driver(self):
        self.driver = webdriver.Chrome(executable_path=self.driverPath, options=webdriver.ChromeOptions())
        self.driver.set_page_load_timeout(20)
        self.driver.set_window_position(800, 1200)
        self.driver.refresh()

    def get_aisle(self, path):

        while self.response is not True:
            try:
                self.driver.get(path)
                WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'results-info')))
                self.response = True
            except NoSuchElementException:
                return None
            except TimeoutException:
                self.driver.close()
                self.init_driver()
                continue
                # self.driver.refresh()
                # time.sleep(random.randint(3, 6))
            except:
                self.init_driver()
                continue

        self.response = False

        soup = BeautifulSoup(self.driver.page_source, 'html.parser')
        try:
            div_tags = soup.find('div', attrs={'class': 'search-tab container'}).find_all('div', attrs={'class': 'tile-in-store-info'})
        except NoSuchElementException:
            div_tags = None

        for div_tag in div_tags:
            str = div_tag.get_text().split('|')

            if len(str) == 2:
                aisle_txt = str[0].split(' ')
                aisle = aisle_txt[1].replace('.','')
                print(aisle)
                return aisle

    def run(self):
        self.init_driver()
        states = self.get_states()
        stores_file = self.readFilePath + '/Stores.xlsx'
        store_wb = load_workbook(stores_file)
        store_sheet = store_wb.active

        target_state = 'TX'

        for i in range(10, len(states)):
            print(states[i])
            if states[i] == target_state:
                store_index = 1
                for store_row in store_sheet:
                    if store_row[6].value == states[i]:
                        if store_index < 25:
                            store_index += 1
                            continue
                        # make a directory for each state
                        dir_name = "../Result/%s" % (states[i])
                        try:
                            # open(os.path.join(os.pardir, "filename"), "w")
                            os.mkdir(dir_name)
                            print("Directory ", dir_name, " Created ")
                        except FileExistsError:
                            print("Directory ", dir_name, " already exists")

                        store_search_url = self.siteUrl + '/store/' + str(store_row[0].value) + '/' + str(
                            store_row[1].value) + '-ar'
                        items_wb = load_workbook("%s/Items.xlsx" % self.readFilePath)
                        items_sheet = items_wb.active

                        # make a new csv file for each store.
                        # print("New File %s-%s.csv created" % (store_row[1].value, store_row[0].value))
                        # file_name = '%s-%s.csv' % (store_row[1].value, store_row[0].value)
                        # with open('../Result/%s' % file_name, mode='w') as f:
                        #     fieldnames = ['Item', 'Aisle']
                        #     writer = csv.DictWriter(f, fieldnames=fieldnames)
                        #     writer.writeheader()
                        #     f.close()
                        #
                        # for item_row in range(1, items_sheet.max_row + 1):
                        #     item = items_sheet.cell(row=item_row, column=1).value
                        #     item_search_url = store_search_url + '/search?query=' + item
                        #     with open('../Result/%s' % file_name, 'w') as f:
                        #         writer = csv.DictWriter(f, fieldnames=fieldnames)
                        #         writer.writerow({'Item': item, 'Aisle': self.get_aisle(item_search_url)})
                        #     f.close()

                        # make a new xlsx file for each store
                        print("File %s-%s.xlsx created" % (store_row[1].value, store_row[0].value))
                        save_path = "%s/%s-%s.xlsx" % (dir_name, store_row[1].value, store_row[0].value)
                        new_wb = Workbook()
                        new_sheet = new_wb.active
                        print(items_sheet.max_row)
                        for item_row in range(1, items_sheet.max_row + 1):
                            item = items_sheet.cell(row=item_row, column=1).value
                            item_search_url = store_search_url + '/search?query=' + item
                            new_sheet.cell(row=item_row, column=1).value = item
                            new_sheet.cell(row=item_row, column=2).value = self.get_aisle(item_search_url)
                            new_wb.save(save_path)


if __name__ == '__main__':
    print("================== starting ==================")
    driverPath = "../driver/chromedriver.exe"
    readFilePath = "../ReadFile"
    writeFilePath = "../result"
    siteUrl = "https://www.walmart.com"
    Scraper = Scraper(driverPath, siteUrl, readFilePath, writeFilePath)
    Scraper.run()

    # https://www.walmart.com/

